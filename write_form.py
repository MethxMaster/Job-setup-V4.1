from openpyxl import load_workbook
import function_support as ex_func

def write_from(first_battery,file_name,output_directory,current_time,form_directory):

	workbook = load_workbook(form_directory + '\\Form\\job setup form.xlsx')

	output_row = [13,15,17,19,21,23,25,27,29,31,33,35,37,39,41,43,45]
	column_row = [6,9,14,35]
	station_list = ['L1_1710','L1_1720','L1_2600','L1_2700','L1_2800','L1_3000',
					'L1_2200','L1_3100','L1_3200','L1_3400','L1_3500','L1_3600',
					'L1_3710','L1_3720','L1_3900','L1_4000','L1_4100']


	workbook['Sheet1'].cell(row=8, column=5).value = current_time.strftime("%d %B %Y")
	workbook['Sheet1'].cell(row=8, column=27).value = 'First battery'

	station_index = 0
	for id_count in output_row :
		if first_battery[station_list[station_index]]['battery'] != '-' :
			workbook['Sheet1'].cell(row=id_count, column=column_row[0]).value = '/'
		station_index += 1

	station_index = 0
	for id_count in output_row :
		if first_battery[station_list[station_index]]['battery'] != '-' :
			if first_battery[station_list[station_index]]['error-message'] == '' :
				workbook['Sheet1'].cell(row=id_count, column=column_row[1]).value = '/'
		station_index += 1


	station_index = 0
	for id_count in output_row :
		workbook['Sheet1'].cell(row=id_count, column=column_row[2]).value = first_battery[station_list[station_index]]['battery']
		station_index += 1

	station_index = 0
	for id_count in output_row :
		if first_battery[station_list[station_index]]['date-time'] != '-' :
			workbook['Sheet1'].cell(row=id_count, column=column_row[3]).value = first_battery[station_list[station_index]]['date-time'][11:19]
		if first_battery[station_list[station_index]]['date-time'] == '-' :
			workbook['Sheet1'].cell(row=id_count, column=column_row[3]).value = '-'
		station_index += 1


	excel_name = 'form-' + file_name + '.xlsx'
	print(output_directory + '/' + excel_name + '.xlsx')####
	workbook.save(filename = output_directory + '/' + excel_name + '.xlsx')
	log_text = 'Excel - excel file has been saved'
	ex_func.notification(log_text)






#if __name__ == '__main__':
#	write_from()